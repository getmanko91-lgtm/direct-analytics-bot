from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.services.analytics_table import AnalyticsRow

HEADERS = (
    "Клиент",
    "Бюджет, мес. ₽",
    "Бюджет, нед. ₽",
    "Расход, ₽",
    "Показы",
    "Клики",
    "CPC, ₽",
    "Цель",
    "Конверсии",
    "CPA, ₽",
    "Примечание",
)

CPA_FILLS = {
    "cpa-high": PatternFill("solid", fgColor="FCE4D6"),
    "cpa-medium": PatternFill("solid", fgColor="FFF2CC"),
    "cpa-low": PatternFill("solid", fgColor="E2EFDA"),
}


def _export_values(row: AnalyticsRow) -> list:
    show = row.show_client_block
    note = row.error or ""
    return [
        row.client_name if show else "",
        row.monthly_budget if show and row.monthly_budget > 0 else None,
        row.weekly_budget if show and row.weekly_budget > 0 else None,
        row.spend if show and not row.error else None,
        row.impressions if show and not row.error else None,
        row.clicks if show and not row.error else None,
        row.cpc if show and row.cpc is not None and not row.error else None,
        row.goal_name,
        row.conversions if not row.error else None,
        row.cpa if row.cpa is not None and not row.error else None,
        note,
    ]


def build_analytics_xlsx(
    rows: list[AnalyticsRow],
    date_from: date,
    date_to: date,
    *,
    cpa_classes: list[str] | None = None,
    vat_percent: int = 22,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    period = date_from.strftime("%d.%m.%Y")
    if date_from != date_to:
        period += f" — {date_to.strftime('%d.%m.%Y')}"

    ws["A1"] = "Сводная аналитика Direct Nikitos Analytics"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период: {period}"
    ws["A3"] = f"Расход и CPA с НДС {vat_percent}%."

    header_row = 5
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    cpa_classes = cpa_classes or []
    for index, row in enumerate(rows):
        excel_row = header_row + 1 + index
        values = _export_values(row)
        for col, value in enumerate(values, start=1):
            ws.cell(row=excel_row, column=col, value=value)

        cpa_class = cpa_classes[index] if index < len(cpa_classes) else ""
        cpa_cell = ws.cell(row=excel_row, column=10)
        if cpa_class in CPA_FILLS:
            cpa_cell.fill = CPA_FILLS[cpa_class]

        for col in (2, 3, 4, 7, 10):
            ws.cell(row=excel_row, column=col).number_format = "#,##0.00"
        for col in (5, 6):
            ws.cell(row=excel_row, column=col).number_format = "#,##0"
        ws.cell(row=excel_row, column=9).number_format = "#,##0.##"

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = len(str(HEADERS[col - 1]))
        for row_idx in range(header_row, header_row + len(rows) + 1):
            value = ws.cell(row=row_idx, column=col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_filename(date_from: date, date_to: date) -> str:
    if date_from == date_to:
        return f"svodka_{date_from.isoformat()}.xlsx"
    return f"svodka_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
