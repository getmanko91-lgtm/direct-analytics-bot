from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.services.client_reports import (
    ClientMonthlyReport,
    WeekMetrics,
    format_money,
    format_period,
    metrics_to_display,
)

FILL_DATE = PatternFill("solid", fgColor="F4B084")
FILL_CONV = PatternFill("solid", fgColor="C6E0B4")
FILL_IMAGE = PatternFill("solid", fgColor="F8CBAD")
FILL_APP = PatternFill("solid", fgColor="FFE699")
FILL_TOTAL = PatternFill("solid", fgColor="D9D9D9")


def export_filename(date_from: date, date_to: date) -> str:
    if date_from == date_to:
        return f"otchety_klientam_{date_from.isoformat()}.xlsx"
    return f"otchety_klientam_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"


def _write_metrics_row(ws, row_idx: int, period_label: str, metrics: WeekMetrics, vat_rate: float) -> None:
    display = metrics_to_display(metrics, vat_rate)
    ws.cell(row=row_idx, column=1, value=period_label)
    ws.cell(row=row_idx, column=2, value=display["conv_spend"])
    ws.cell(row=row_idx, column=3, value=display["conv_count"])
    ws.cell(row=row_idx, column=4, value=display["conv_price"])
    ws.cell(row=row_idx, column=5, value=display["image_spend"])
    ws.cell(row=row_idx, column=6, value=display["image_impressions"])
    ws.cell(row=row_idx, column=7, value=display["image_cpm"])
    ws.cell(row=row_idx, column=8, value=display["image_conversions"])
    ws.cell(row=row_idx, column=9, value=display["app_spend"])
    ws.cell(row=row_idx, column=10, value=display["app_installs"])
    ws.cell(row=row_idx, column=11, value=display["app_cpi"])
    ws.cell(row=row_idx, column=12, value=display["app_revenue"])
    ws.cell(row=row_idx, column=13, value=display["total_spend"])


def _write_report_header(ws, start_row: int) -> int:
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)
    cell = ws.cell(row=start_row, column=1, value="Дата")
    cell.font = header_font
    cell.fill = FILL_DATE
    cell.alignment = center

    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=4)
    cell = ws.cell(row=start_row, column=2, value="реклама на конверсии")
    cell.font = header_font
    cell.fill = FILL_CONV
    cell.alignment = center

    ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=8)
    cell = ws.cell(row=start_row, column=5, value="реклама на имидж")
    cell.font = header_font
    cell.fill = FILL_IMAGE
    cell.alignment = center

    ws.merge_cells(start_row=start_row, start_column=9, end_row=start_row, end_column=12)
    cell = ws.cell(row=start_row, column=9, value="реклама приложения")
    cell.font = header_font
    cell.fill = FILL_APP
    cell.alignment = center

    ws.merge_cells(start_row=start_row, start_column=13, end_row=start_row + 1, end_column=13)
    cell = ws.cell(row=start_row, column=13, value="расход общий")
    cell.font = header_font
    cell.fill = FILL_TOTAL
    cell.alignment = center

    subheaders = [
        (2, "расход", FILL_CONV),
        (3, "конверсии", FILL_CONV),
        (4, "цена конверсии", FILL_CONV),
        (5, "расход", FILL_IMAGE),
        (6, "показы", FILL_IMAGE),
        (7, "CPM (стоимость 1000 показов)", FILL_IMAGE),
        (8, "конверсии", FILL_IMAGE),
        (9, "расход", FILL_APP),
        (10, "установка", FILL_APP),
        (11, "цена установки", FILL_APP),
        (12, "доход", FILL_APP),
    ]
    sub_row = start_row + 1
    for col, title, fill in subheaders:
        cell = ws.cell(row=sub_row, column=col, value=title)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center

    return start_row + 2


def _write_client_section(ws, start_row: int, report: ClientMonthlyReport, vat_rate: float) -> int:
    ws.cell(row=start_row, column=1, value=report.client_name).font = Font(bold=True, size=12)
    row = start_row + 1

    if report.error:
        ws.cell(row=row, column=1, value=report.error)
        return row + 2

    row = _write_report_header(ws, row)

    for (week_from, week_to), metrics in zip(report.weeks, report.week_metrics, strict=True):
        _write_metrics_row(ws, row, format_period(week_from, week_to), metrics, vat_rate)
        row += 1

    total_label = f"Итого {format_period(report.weeks[0][0], report.weeks[-1][1])}" if report.weeks else "Итого"
    _write_metrics_row(ws, row, total_label, report.total, vat_rate)
    row += 1

    ws.cell(row=row, column=11, value="ПЛАН").font = Font(bold=True)
    if report.plan_budget > 0:
        ws.cell(row=row, column=13, value=format_money(report.plan_budget))
    row += 2
    return row


def build_client_reports_xlsx(
    reports: list[ClientMonthlyReport],
    date_from: date,
    date_to: date,
    *,
    vat_percent: int = 22,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчеты клиентам"

    period = date_from.strftime("%d.%m.%Y")
    if date_from != date_to:
        period += f" — {date_to.strftime('%d.%m.%Y')}"

    ws["A1"] = "Отчеты клиентам — Direct Nikitos Analytics"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период: {period}"
    ws["A3"] = f"Расход с НДС {vat_percent}%. Цена конверсии / CPM / цена установки — с НДС."

    vat_rate = vat_percent / 100
    row = 5
    for report in reports:
        row = _write_client_section(ws, row, report, vat_rate)

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
